#!/usr/bin/env python
# coding: utf-8

# # 舞光 LED 產品型錄 RAG 系統 v3.2 (2026-05-13)
# 
# ## 本版本與 v3.1 差異
# 1. **Chunk metadata 新增 `image_paths`**: 把每頁的圖片實體路徑 (PNG) 寫進 chunk meta；保留作為後續實驗的 hook，目前 step 10b 不使用。
# 2. **步驟 10b 維持文字輸入** — 多模態 (gpt-4o + image_url) 實驗結果 acc_any 41.2%→17.6% 退步，已撤回。gpt-5.4 多模態 acc_any 35.3% 仍低於 v3.1 text-only baseline (41.2%)，亦撤回。
# 
# ## 系統目標
# 從 388 頁的舞光 (Dancelight) LED 產品型錄 PDF 建立檢索式問答系統。給定客戶用自然語言或標案規格描述的需求（例：「LED 投射燈，≦50W，6000K，≧6000lm，IP65」），系統需要從型錄中找出最匹配的產品型號（如 `E-FLCS50D`）並以「★ 推薦 + 4 備選」格式回覆給使用者。
# 
# ## Pipeline 總覽
# 
# ```
# PDF (388 pages)
#    ├─► PyMuPDF 文字擷取 (步驟 1) ──┐
#    └─► opendataloader 結構化解析 (步驟 2) ─► 圖片清單
#                                               │
#                                               ▼
#                             minicpm-v 視覺描述 (步驟 3，含快取)
#                                               │
#                                               ▼
#    ┌──────────────── 結構化 Chunking (步驟 4) ─────────────┐
#    │  每頁 1 個 chunk = 結構化 metadata + 摘要 + 原文 + 圖描述  │
#    │  metadata: page / series_name / category / models /     │
#    │            wattages / color_temps / lumens / IP /        │
#    │            features                                       │
#    └────────────────────────────────────────────────────────────┘
#                                               │
#                           ┌───────────────────┼────────────────────┐
#                           ▼                   ▼                    ▼
#               步驟 5: BM25 索引       步驟 6: BGE-M3 向量化     步驟 7-8: 函式
#               (jieba + 燈具詞典)      (語意檢索, 快取 .npy)     decompose / metadata_filter
#                                                                   / hybrid_retrieve
#                                               │
#                           ┌───────────────────┴────────────────────┐
#                           │     步驟 10: 檢索 + Reranker pipeline    │
#                           │  ① decompose_query  (regex → specs)    │
#                           │  ② metadata_filter  (合格 chunk index)   │
#                           │  ③ expand_specs / add_synonyms 重寫查詢  │
#                           │  ④ hybrid_retrieve (BM25+vec, top-50)   │
#                           │  ⑤ BGE-reranker     (CrossEncoder, top-20) │
#                           └────────────────────────────────────────────┘
#                                               │
#                                               ▼
#                 ┌─────────── 步驟 10b: LLM 答案生成 ──────────┐
#                 │  gpt-oss:120b (Ollama) 從 top-20 產生：       │
#                 │    ★ 推薦：<型號> — <規格>                    │
#                 │    備選 1～4：<型號> — <規格>                  │
#                 │  Context 上限 12000 字、num_ctx=8192          │
#                 └──────────────────────────────────────────────────┘
#                                               │
#                                               ▼
#               步驟 11: xlsx 輸出 (LLM 答案 + top-20)
#                                               │
#                                               ▼
#               步驟 12: 命中率統計
#                   ┌── 檢索層: gold 在 top-20 chunks (寬鬆)
#                   ├── LLM acc_any: gold 在 5 選項 (中)
#                   └── LLM acc_rec: gold 在 ★ 推薦那行 (嚴格)
# ```
# 
# ## 為什麼要這樣設計？
# 
# 1. **結構化 chunking**: 直接用整頁文字做 chunk 會讓檢索無法依規格條件過濾。我們在 chunking 階段就用 regex 把每頁的型號、瓦數、色溫、光通量、IP 抽出來放進 metadata，後續即可用結構化條件做 metadata filter。
# 2. **混合檢索 (BM25 + Vector)**:
#    - **BM25**：對精確字詞（型號 `E-FLCS50D`、規格 `IP65`）有絕對優勢
#    - **BGE-M3 向量**：對同義語意（「投射燈」≈「泛光燈」）和模糊查詢有優勢
#    - 兩者各 0.5 權重融合，min-max 正規化後相加
# 3. **Metadata 加分機制**: 過濾後合格的 chunk 乘 1.3，不合格乘 0.7。不是硬過濾——避免 metadata 抽取錯誤時把正解直接踢掉。
# 4. **同義詞群組 (SYNONYMS)**: 型錄使用「泛光燈」，但客戶常說「投射燈/探照燈/聚光燈」。同一群組的詞彼此都認，避免術語差異造成 miss。
# 5. **BGE-reranker (CrossEncoder)**: 第二階段精排。前面 hybrid 取 50 個候選，reranker 一對對重新評分後挑前 20，補正 BM25 + 向量都漏掉的細節。
# 6. **5 選項 LLM 答案 (★ 推薦 + 4 備選)**: 不是直接吐單一答案。強迫 LLM 從 20 候選收斂到 1 推薦 + 4 備選，給業務 fallback 空間。也讓我們可同時用嚴格 (acc_rec) 和寬鬆 (acc_any) 兩個指標衡量品質。
# 
# ## 關鍵超參數
# 
# | 參數 | 值 | 說明 |
# |---|---|---|
# | `TOP_K` | 20 | 最終餵給 LLM 的 chunk 數 |
# | `RETRIEVE_K` | 50 | 第一階段檢索保留候選給 reranker |
# | `BM25_WEIGHT` | 0.5 | BM25 在 hybrid score 的權重 |
# | `VECTOR_WEIGHT` | 0.5 | BGE-M3 向量在 hybrid score 的權重 |
# | metadata 加分 | 1.3 / 0.7 | 符合/不符合 metadata 條件的乘數 |
# | `CTX_MAX_CHARS` | 12000 | LLM 答案生成時的 context 上限字元數 |
# | `num_ctx` | 8192 | gpt-oss:120b 的 context window |
# | `num_predict` | 1600 | LLM 輸出 token 上限 (5 選項通常 < 800) |
# 
# ## 模型選用
# 
# | 用途 | 模型 | 備註 |
# |---|---|---|
# | Embedding | `BAAI/bge-m3` (1024-dim) | 中文最佳開源 embedder，可離線執行 |
# | Reranker | `BAAI/bge-reranker-v2-m3` | 對應 BGE-M3 的精排器 |
# | **答案 LLM** | `gpt-oss:120b` (Ollama) | **本地 5 選項生成**；同時兼任 HyDE |
# | 視覺描述 | `minicpm-v` (Ollama) | 圖片轉文字描述，結果含入 chunk |
# 
# > 全 pipeline 可離線執行（不打雲端 API）。如要切換答案 LLM 為 OpenAI gpt-5 / Gemini，只需替換 `run_llm()` 內部呼叫，prompt 模板不動。
# 
# ## 輸入/輸出檔案
# 
# - **輸入**:
#   - `2025舞光LED21st(單頁水印可搜尋).pdf` — 型錄
#   - `output_opendataloader/*.json` — 結構化解析結果
#   - `Training.xlsx` — 測試題庫 (詢問問題 / 期望回答 兩欄)
#   - `img_descriptions_cache.json` — 圖片描述快取
# - **輸出**:
#   - `bge_m3_embeddings/chunk_embeddings.npy` — 預計算的 chunk 向量
#   - `rag_results.xlsx` — 每題：LLM 答案 + top-20 chunks + 命中率統計
#   - `results` (記憶體中): list[dict]，每筆含 `qid / original / expected / specs / docs / answer / llm_time / llm_error`
# 
# ## 評估指標
# 
# 三個指標分層評估：
# 
# | 指標 | 算法 | 用途 |
# |---|---|---|
# | 檢索命中率 | gold 型號**任一**出現在 top-20 chunk 文字裡 | 看 retrieval+rerank 階段是否漏 |
# | LLM acc_any | gold 型號出現在 5 選項中**任一**位 | 看 LLM 是否從 20 候選中認出正解 |
# | LLM acc_rec | gold 型號出現在 **★ 推薦** 那行 | 看 LLM 是否能把正解放第一順位 |
# 
# 理想差距：檢索 ≈ acc_any ≈ acc_rec（系統前後段都對齊）。實務上：
# - 檢索 > acc_any → LLM 看到正解但沒選 (prompt/context 問題)
# - acc_any > acc_rec → LLM 認出正解但放在備選 (排序信心不足)

# ## 步驟 0：依賴套件與設定
# 
# 匯入所有套件並定義全域常數。重要設定：
# 
# | 變數 | 用途 |
# |---|---|
# | `PDF_PATH` | 型錄 PDF 路徑 |
# | `ODL_JSON` | opendataloader 解析後的結構化 JSON |
# | `TRAINING_XLSX` | 測試題庫 (詢問問題 / 期望回答) |
# | `EMBED_DIR` | 預計算 embedding 存放位置 (`.npy` 快取) |
# | `LOCAL_MODEL` | HyDE 用的本地 LLM (`gpt-oss:120b` via Ollama) |
# | `EMBED_MODEL` | `BAAI/bge-m3` — 中文向量模型 |
# | `RERANK_MODEL` | `BAAI/bge-reranker-v2-m3` — CrossEncoder 精排器 |
# | `VISION_MODEL` | `minicpm-v` — 圖片描述模型 |
# | `MIN_IMG_SIZE` | 30KB；過小的圖通常是 icon/logo，跳過視覺處理省時間 |
# | `TOP_K` / `RETRIEVE_K` | 最終 20，第一階段保留 50 給 reranker |
# | `BM25_WEIGHT` / `VECTOR_WEIGHT` | hybrid 融合權重，各 0.5 |

# In[3]:


# Cell 2 — 匯入套件 + 全域常數設定 (路徑、模型名稱、超參數)
import json, os, re, sys, time, base64, glob, signal, shutil
import fitz  # PyMuPDF
import numpy as np
import pandas as pd
import jieba
import ollama
from rank_bm25 import BM25Okapi
from sentence_transformers import SentenceTransformer, CrossEncoder
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from dotenv import load_dotenv
from openai import OpenAI

load_dotenv('.env')
openai_client = OpenAI()

# ── 設定 ──
PDF_PATH = "./2025舞光LED21st(單頁水印可搜尋).pdf"
ODL_JSON = "./output_opendataloader/2025舞光LED21st(單頁水印可搜尋).json"
ODL_IMG_DIR = "./output_opendataloader/2025舞光LED21st(單頁水印可搜尋)_images"
QUESTION_XLSX = "./question.xlsx"
ANSWER_XLSX   = "./answer.xlsx"
CHROMA_DIR = "./chroma_rag_v3"
RESULTS_XLSX = "./rag_results.xlsx"

LOCAL_MODEL = "gemma4:31b"
CLOUD_LLM = "qwen3.6:latest"  # v3.3 試: Ollama 本地 qwen3.6 作為答案 LLM
EMBED_MODEL = "Qwen/Qwen3-Embedding-8B"  # v3.3 試: 改用 Qwen3-Embedding-8B (4096-dim, fp16)
EMBED_KWARGS = {"device": "cuda", "model_kwargs": {"torch_dtype": "float16"}}
RERANK_MODEL = "BAAI/bge-reranker-v2-m3"
VISION_MODEL = "minicpm-v"
MIN_IMG_SIZE = 30_000
TOP_K = 20
RETRIEVE_K = 50
BM25_WEIGHT = 0.5
VECTOR_WEIGHT = 0.5
EMBED_DIR = "./qwen3_8b_embeddings"  # v3.3: 用 Qwen3-Embedding-8B 的 cache

print("=" * 60)
print("舞光 LED 產品型錄 RAG 系統 v3.1")
print("  BGE-M3 + BGE-reranker + 結構化 Chunking")
print("=" * 60)


# ## 步驟 1：PyMuPDF 文字擷取
# 
# 用 PyMuPDF (`fitz`) 讀整本 PDF，逐頁抽取純文字。
# 
# **為什麼用 PyMuPDF**: 速度快、保留閱讀順序、能處理水印 PDF。388 頁完整跑完約 1 秒。
# 
# **輸出**: `pymupdf_texts: dict[int, str]` ─ key 是 1-based 頁碼，value 是該頁的純文字。空白頁會被跳過。
# 
# **侷限**: 純文字抽取會遺失：表格結構、圖片、版面位置。表格仍會被抽成文字但欄位順序不一定對 ─ 這就需要下一步的 opendataloader 來補。

# In[4]:


# Cell 4 — 用 PyMuPDF 對 PDF 逐頁抽純文字，輸出 pymupdf_texts: {page_num: text}
print("\n" + "=" * 60)
print("步驟 1: PyMuPDF 文字擷取")
print("=" * 60)

doc = fitz.open(PDF_PATH)
N_PAGES = len(doc)
pymupdf_texts = {}
for i in range(N_PAGES):
    text = doc[i].get_text().strip()
    if text:
        pymupdf_texts[i + 1] = text
doc.close()
print(f"有文字的頁面: {len(pymupdf_texts)} / {N_PAGES}")


# ## 步驟 2：opendataloader 結構化解析
# 
# 讀取 opendataloader 輸出的 JSON，這個工具把 PDF 拆成結構化的元素：表格、段落、**圖片** (含檔名與頁碼)。
# 
# **用途**: 我們主要拿這份 JSON 的圖片清單 ─ PyMuPDF 不會告訴你哪頁有什麼圖、檔名是什麼，但 opendataloader 會。
# 
# **過濾**: 只保留 `>= MIN_IMG_SIZE` (30KB) 的圖片。小圖通常是 icon、logo、佈景元素，丟給視覺模型只會浪費時間。
# 
# **輸出**:
# - `odl_images: dict[int, list[str]]` — 每頁的圖片檔案路徑列表
# - `odl_texts: dict[int, list[(type, content)]]` — 結構化文字片段（目前未使用，但保留以備擴展）

# In[5]:


# Cell 6 — 讀 opendataloader JSON，收集每頁的圖片檔案清單 (>= 30KB 才保留)
print("\n" + "=" * 60)
print("步驟 2: opendataloader 結構化資料")
print("=" * 60)

with open(ODL_JSON, "r") as f:
    odl_data = json.load(f)

odl_texts = {}
odl_images = {}
for kid in odl_data["kids"]:
    pn = kid.get("page number", 0)
    kt = kid.get("type", "")
    if kt == "image":
        src = kid.get("source", "")
        if src:
            fp = os.path.join("./output_opendataloader", src)
            if os.path.exists(fp) and os.path.getsize(fp) >= MIN_IMG_SIZE:
                odl_images.setdefault(pn, []).append(fp)
    else:
        c = (kid.get("content", "") or "").strip()
        if c:
            odl_texts.setdefault(pn, []).append((kt, c))

print(f"有結構化文字的頁面: {len(odl_texts)}")


# ## 步驟 3：minicpm-v 圖片描述（含快取）
# 
# 對步驟 2 收集到的每張圖片呼叫 `minicpm-v` 視覺模型，生成 50–100 字的中文描述。描述會記錄：
# 1. 圖片上的角度數字（光束角、旋轉角）
# 2. 產品型號（型錄圖通常會印型號）
# 3. 尺寸標註
# 4. 外觀特徵（形狀、顏色、安裝方式）
# 
# **為什麼要做**: 純文字 chunk 無法搜「附鏡面鋁板」「樓梯燈」這種純從圖片才看得出來的特徵。把視覺資訊轉文字後串在 chunk 裡一起索引，BM25 / 向量都能利用。
# 
# **快取機制**: `img_descriptions_cache.json` 以圖片路徑為 key 儲存描述。已處理過的圖直接讀快取，新圖才呼叫模型。每處理 20 張存檔一次，避免中途斷電白工。
# 
# **逾時保護**: 每張圖最多 30 秒，避免 minicpm-v 偶爾卡死拖慢整體流程。

# In[ ]:


# Cell 8 — 對所有圖片呼叫 minicpm-v 生成中文描述（含 path-keyed 磁碟快取與 30s 逾時保護）
print("\n" + "=" * 60)
print("步驟 3: minicpm-v 圖片描述")
print("=" * 60)

VISION_PROMPT = """請用繁體中文簡要描述這張燈具產品圖片的重要資訊，特別注意：
1. 圖片上標註的角度數字（如旋轉角度、光束角）
2. 產品型號文字
3. 尺寸標註
4. 產品的外觀特徵（形狀、顏色、安裝方式）
請只描述圖片中看到的內容。用50-100字簡述。"""

IMG_CACHE = "./img_descriptions_cache.json"
if os.path.exists(IMG_CACHE):
    with open(IMG_CACHE, "r") as f:
        img_cache = json.load(f)
    print(f"載入圖片描述快取: {len(img_cache)} 筆")
else:
    img_cache = {}

page_img_descs = {}
processed = 0
skipped = 0

for pn in sorted(odl_images.keys()):
    descs = []
    for fp in odl_images[pn]:
        fn = os.path.basename(fp)
        if fp in img_cache:
            descs.append(f"[圖片{fn}] {img_cache[fp]}")
            skipped += 1
            continue
        processed += 1
        print(f"  [{processed}] 第{pn}頁 {fn}...", end=" ", flush=True)
        with open(fp, "rb") as f:
            img_b64 = base64.b64encode(f.read()).decode()
        t0 = time.time()
        try:
            def _timeout_handler(signum, frame):
                raise TimeoutError()
            old_handler = signal.signal(signal.SIGALRM, _timeout_handler)
            signal.alarm(30)
            resp = ollama.chat(model=VISION_MODEL, messages=[{
                "role": "user", "content": VISION_PROMPT, "images": [img_b64],
            }])
            signal.alarm(0)
            signal.signal(signal.SIGALRM, old_handler)
            desc = resp["message"]["content"].strip()
        except TimeoutError:
            signal.alarm(0)
            desc = "(超時)"
            print("TIMEOUT", end=" ")
        except Exception as e:
            signal.alarm(0)
            desc = f"(失敗)"
        img_cache[fp] = desc
        descs.append(f"[圖片{fn}] {desc}")
        print(f"({time.time()-t0:.1f}s)")
        if processed % 20 == 0:
            with open(IMG_CACHE, "w") as f:
                json.dump(img_cache, f, ensure_ascii=False)
    if descs:
        page_img_descs[pn] = descs

with open(IMG_CACHE, "w") as f:
    json.dump(img_cache, f, ensure_ascii=False)
print(f"新處理: {processed}, 快取命中: {skipped}")


# ## 步驟 4：結構化 Chunking — 每頁一個 chunk
# 
# 這是整個系統的核心。不像一般 RAG 用固定字數切 chunk，我們**一頁一個 chunk** 並抽取結構化 metadata，因為型錄每頁就是一個產品系列，跨頁切割會破壞語意。
# 
# ### `extract_products_from_page(page_num, text)` 抽取流程
# 
# 1. **找型號**: regex `[A-Z][A-Z0-9-]{3,}[A-Z0-9]` ─ 找「大寫開頭、4+ 字元、英數混合」的字串作為候選型號。
# 2. **找系列名**: 取「含『燈』字、3–25 字元、不是規格表頭」的短行，第一個就當系列名稱。
# 3. **抽規格** (regex):
#    - 瓦數 `(\d+)W`
#    - 色溫 `(\d+)K`
#    - 光通量 `(\d+)\s*LM`
#    - IP 等級 `IP\s*(\d+)`
#    - 電壓 / 發光角度
# 4. **判類別**: 用關鍵字字典 `category_keywords` 比對。優先看系列名，找不到再看全文。同義詞分群（投射燈/泛光燈/探照燈/聚光燈 都歸 `泛光燈`）。
# 5. **抽特色**: 節能標章 / 感應 / 調光 / 防眩 / 防水 / 全電壓 任一出現就記。
# 
# ### Chunk 結構
# 
# 每個 chunk 包含三段：
# ```
# 【產品】系列名
# 【類別】類別
# 【型號】M1, M2, M3, ...
# 【瓦數】50W
# 【色溫】6000K
# 【光通量】6500LM
# 【IP防護】IP65
# 【特色】節能標章, 防水
# 
# --- 原始內容 (第N頁) ---
# [整頁原始文字]
# 
# [圖片描述 1]
# [圖片描述 2]
# ...
# ```
# 
# 頭部結構化摘要讓 BM25/向量在搜「IP65」「6000K」這類規格時有更高的 token 集中度。底部原文保證資訊完整。圖片描述讓視覺特徵也可被檢索。
# 
# ### Metadata 欄位（附在每個 chunk 上）
# 
# `page` / `series_name` / `category` / `models` / `wattages` / `color_temps` / `lumens` / `ip_rating` / `features` — 這些是後續 `metadata_filter` 過濾的依據。

# In[ ]:


# Cell 10 — extract_products_from_page() 用 regex 抽型號/規格/類別，建立含 metadata 的 chunk
print("\n" + "=" * 60)
print("步驟 4: 結構化 Chunking (每產品系列)")
print("=" * 60)


def extract_products_from_page(page_num, text):
    """從頁面文字中提取產品系列，每個系列含型號和規格"""
    products = []

    # 找所有型號
    model_pattern = r'[A-Z][A-Z0-9-]{3,}[A-Z0-9](?:\b|$)'
    all_models = list(dict.fromkeys(re.findall(model_pattern, text)))

    if not all_models:
        return products

    # 找產品系列名稱 (含「燈」的短行)
    lines = text.split("\n")
    product_names = []
    for line in lines:
        line = line.strip()
        if "燈" in line and 3 < len(line) < 25:
            if not any(kw in line for kw in ["產品型號", "消耗", "輸入", "材質", "色溫", "光通量", "型錄"]):
                product_names.append(line)

    series_name = product_names[0] if product_names else f"第{page_num}頁產品"

    # 提取規格值
    wattages = list(dict.fromkeys(re.findall(r'(\d+)W\b', text)))
    color_temps = list(dict.fromkeys(re.findall(r'(\d+)K\b', text)))
    lumens = list(dict.fromkeys(re.findall(r'(\d+)\s*LM\b', text, re.I)))
    ip_match = re.findall(r'IP\s*(\d+)', text)
    ip_rating = ip_match[0] if ip_match else ""
    voltage_match = re.findall(r'(\d+-\d+V|\d+V)', text)
    voltage = voltage_match[0] if voltage_match else ""
    beam_match = re.findall(r'(\d+)°', text)
    beam_angle = beam_match[0] if beam_match else ""

    # 檢查特殊功能
    features = []
    if "節能標章" in text: features.append("節能標章")
    if "感應" in text: features.append("感應")
    if "調光" in text: features.append("調光")
    if "防眩" in text: features.append("防眩")
    if "防水" in text: features.append("防水")
    if "全電壓" in text: features.append("全電壓")

    # 判斷產品類別 (優先看產品名稱，再看全文)
    category = ""
    category_keywords = [
    ("泛光燈", ["泛光燈", "投射燈", "探照燈", "聚光燈"]),
    ("輕鋼架燈", ["輕鋼架燈", "T-BAR燈", "格柵燈", "嵌入式日光燈", "辦公室燈"]),
    ("工事燈", ["工事燈", "吊管燈", "吊管式吊燈", "停車場燈"]),
    ("中東燈", ["中東燈", "吊管燈"]),
    ("崁燈", ["崁燈", "筒燈", "嵌燈", "下照燈", "downlight"]),
    ("步道燈", ["步道燈", "車道燈", "階梯燈", "引導燈", "夜間指引燈"]),
    ("吸頂燈", ["吸頂燈", "環形燈", "雲朵燈"]),
    ("平板燈", ["平板燈", "面板燈", "LED面板"]),
    ("支架燈", ["支架燈", "層板燈", "T8支架燈"]),
    ("防潮燈", ["防潮燈", "浴室燈", "防水吸頂燈"]),
    ("軌道燈", ["軌道燈"]),
    ("吊燈", ["吊燈"]),
    ("壁燈", ["壁燈"]),
    ("路燈", ["路燈"]),
    ("高燈", ["高燈", "景觀燈"]),
    ("洗牆燈", ["洗牆燈"]),
    ("燈管", ["燈管"]),
    ("燈泡", ["燈泡"]),
]
    # 先看產品名稱 (更精確)
    name_text = series_name + " " + " ".join(product_names)
    for cat, kws in category_keywords:
        if any(kw in name_text for kw in kws):
            category = cat
            break
    # 若產品名稱無法判斷，再看全文
    if not category:
        for cat, kws in category_keywords:
            if any(kw in text for kw in kws):
                category = cat
                break

    # 建立結構化 metadata
    meta = {
        "page": page_num,
        "series_name": series_name,
        "category": category,
        "models": ",".join(all_models[:10]),
        "wattages": ",".join(wattages[:5]),
        "color_temps": ",".join(color_temps[:5]),
        "lumens": ",".join(lumens[:5]),
        "ip_rating": ip_rating,
        "features": ",".join(features),
        # v3.2: 把實體圖片路徑也帶進 metadata，供步驟 10b 多模態 LLM 使用
        "image_paths": list(odl_images.get(page_num, [])),
    }

    # 建立 chunk 文字 (結構化摘要 + 原始文字)
    summary_parts = [
        f"【產品】{series_name}",
        f"【類別】{category}" if category else "",
        f"【型號】{', '.join(all_models[:8])}",
        f"【瓦數】{', '.join(wattages[:5])}W" if wattages else "",
        f"【色溫】{', '.join(color_temps[:5])}K" if color_temps else "",
        f"【光通量】{', '.join(lumens[:5])}LM" if lumens else "",
        f"【IP防護】IP{ip_rating}" if ip_rating else "",
        f"【電壓】{voltage}" if voltage else "",
        f"【發光角度】{beam_angle}°" if beam_angle else "",
        f"【特色】{', '.join(features)}" if features else "",
    ]
    summary = "\n".join(p for p in summary_parts if p)

    # 加入圖片描述
    img_desc = ""
    if page_num in page_img_descs:
        img_desc = "\n" + "\n".join(page_img_descs[page_num])

    # 完整 chunk = 結構化摘要 + 原始文字 + 圖片描述
    chunk_text = f"{summary}\n\n--- 原始內容 (第{page_num}頁) ---\n{text}{img_desc}"

    products.append({"text": chunk_text, "metadata": meta})
    return products


# 建立產品 chunks
chunks = []
chunk_metas = []

for pn in sorted(pymupdf_texts.keys()):
    text = pymupdf_texts[pn]
    products = extract_products_from_page(pn, text)
    if products:
        for p in products:
            chunks.append(p["text"])
            chunk_metas.append(p["metadata"])
    else:
        # 非產品頁面也保留 (目錄、索引等)
        chunks.append(f"[第{pn}頁]\n{text[:1000]}")
        chunk_metas.append({"page": pn, "series_name": "", "category": "",
                           "models": "", "wattages": "", "color_temps": "",
                           "lumens": "", "ip_rating": "", "features": "",
                           "image_paths": list(odl_images.get(pn, []))})

print(f"共建立 {len(chunks)} 個 Chunk")
# 統計有產品資訊的 chunk
has_models = sum(1 for m in chunk_metas if m.get("models"))
has_category = sum(1 for m in chunk_metas if m.get("category"))
print(f"  含產品型號: {has_models}")
print(f"  含產品類別: {has_category}")

# 顯示幾個範例
for pn in [103, 189, 239]:
    for i, m in enumerate(chunk_metas):
        if m.get("page") == pn and m.get("models"):
            print(f"\n  範例 [p.{pn}]: {m['series_name']} | {m['category']} | models={m['models'][:40]}")
            print(f"    W={m['wattages']} K={m['color_temps']} LM={m['lumens']} IP={m['ip_rating']}")
            break


# ## 步驟 4.5：載入產品類別同義詞表 (category_synonyms.xlsx)
# 
# **為什麼新增這個步驟？**
# 之前 SYNONYMS / cat_map / lamp_terms 是寫死在程式碼裡 (上百行硬編碼 dict)。改成從 `category_synonyms.xlsx` 動態載入後，業務同事只需要編輯試算表就能調整同義詞行為，不用改 code。
# 
# **載入內容**
# - **Sheet 1 「產品類別同義詞」** (18 個正式名稱 + 各自的 1〜5 個別名)
#   - 用 union-find 合併「有交集」的群（例：「吊管燈」同時出現在工事燈、中東燈兩列 → 合併成同一個同義詞群）
#   - 產出 `ALIAS_TO_CANON` (別名 → 型錄正式名稱)
#   - 產出 `NAME_TO_GROUP` / `SYN_GROUPS_LIST` (任一名詞 → 整個同義詞群)
# - **Sheet 2 「產品類別階層」** (大類 → 中類 → 小類)
#   - 產出 `SIBLINGS_MAP` (canonical → 同中類兄弟列表)，可作為未來的軟同類擴展
# 
# **LAMP_TERMS** = 所有正式名 + 別名 + 專業名詞，按長度倒序，餵給 jieba 確保「T-BAR燈」「downlight」「T8支架燈」等複合詞不被切碎。
# 
# **下游用法**
# | 步驟 | 用到的全域 | 作用 |
# |---|---|---|
# | 步驟 5 BM25 | `LAMP_TERMS` | 加入 jieba 詞典 |
# | 步驟 7 decompose_query | `ALIAS_TO_CANON` | 將查詢中的別名（投射燈）對到 canonical（泛光燈） |
# | 步驟 7 metadata_filter | `NAME_TO_GROUP` | chunk metadata category 用整個同義詞群比對 |
# | 步驟 8 add_synonyms | `NAME_TO_GROUP` | BM25 + 向量 query 兩端皆做同義詞展開 |
# 

# In[ ]:


# Cell — 步驟 4.5: 載入 category_synonyms.xlsx → LAMP_TERMS / ALIAS_TO_CANON / NAME_TO_GROUP / SIBLINGS_MAP
print("\n" + "=" * 60)
print("步驟 4.5: 產品類別同義詞表 (category_synonyms.xlsx)")
print("=" * 60)

SYN_XLSX_PATH = "category_synonyms.xlsx"
if not os.path.exists(SYN_XLSX_PATH):
    SYN_XLSX_PATH = "../category_synonyms.xlsx"
print(f"來源檔案: {SYN_XLSX_PATH}")

# Sheet 1：產品類別同義詞 → 建立 raw_groups, 然後 union-find 合併有交集的群
_df_syn = pd.read_excel(SYN_XLSX_PATH, sheet_name="產品類別同義詞")

def _is_canon_row(name):
    return isinstance(name, str) and (name.endswith("燈") or name in ("燈管", "燈泡"))

_raw_groups = []
ALIAS_TO_CANON = {}
for _, _row in _df_syn.iterrows():
    _canon = _row.iloc[0]
    if not _is_canon_row(_canon):
        continue
    _syns = [str(s).strip() for s in _row.iloc[1:].tolist()
             if isinstance(s, str) and s.strip()]
    _raw_groups.append({_canon, *_syns})
    ALIAS_TO_CANON.setdefault(_canon, _canon)
    for _s in _syns:
        ALIAS_TO_CANON.setdefault(_s, _canon)

def _merge_groups(gs):
    """Union-find: 將有交集的群合併"""
    changed = True
    while changed:
        changed = False
        for i in range(len(gs)):
            for j in range(i + 1, len(gs)):
                if gs[i] & gs[j]:
                    gs[i] = gs[i] | gs[j]
                    gs.pop(j)
                    changed = True
                    break
            if changed:
                break
    return gs

SYN_GROUPS_LIST = _merge_groups(_raw_groups)

# 任一名稱 → 整個同義詞群
NAME_TO_GROUP = {}
for _g in SYN_GROUPS_LIST:
    for _term in _g:
        NAME_TO_GROUP[_term] = _g

# LAMP_TERMS = 所有別名 + 正式名 + 專業名詞 (長詞優先，對 jieba/regex 都安全)
_PRO_TERMS = ["色溫", "光通量", "演色性", "防水", "節能標章", "全電壓",
              "感應型", "調光", "防眩", "崁入孔", "發光角度", "光束角",
              "燈座", "燈具", "燈殼", "日光燈"]
LAMP_TERMS = sorted(set(list(ALIAS_TO_CANON.keys()) + _PRO_TERMS),
                    key=lambda x: -len(x))

# Sheet 2：產品類別階層 → SIBLINGS_MAP (中類兄弟)
_df_hier = pd.read_excel(SYN_XLSX_PATH, sheet_name="產品類別階層")
SIBLINGS_MAP = {}
_mid_col = _df_hier.columns[1]
_small_col = _df_hier.columns[2]
for _mid, _grp in _df_hier.groupby(_mid_col):
    if not isinstance(_mid, str) or "說明" in _mid:
        continue
    _sibs = [s for s in _grp[_small_col].tolist() if isinstance(s, str)]
    for _s in _sibs:
        SIBLINGS_MAP[_s] = _sibs

print(f"  同義詞群: {len(SYN_GROUPS_LIST)} 群 (union-find 合併後)")
print(f"  ALIAS_TO_CANON: {len(ALIAS_TO_CANON)} 個別名→正式名")
print(f"  LAMP_TERMS (jieba 詞典): {len(LAMP_TERMS)} 詞")
print(f"  SIBLINGS_MAP (中類兄弟): {len(SIBLINGS_MAP)} 個 canonical")
print(f"  範例同義詞群:")
for _g in SYN_GROUPS_LIST[:4]:
    print(f"    {sorted(_g)}")


# ## 步驟 4.7：產品註解獨立資料庫（chunk annotation sidecar DB）
# 
# **動機**: BM25 / BGE-M3 / reranker 三層都是看 chunk 原文。型錄原文常常一頁有 5-10 個型號、規格散在表格不同欄位，
# 單純語意比對容易把客戶問句對到附近但錯誤的產品。我們對每個 chunk 用 LLM 預先生成一段「結構化註解」
# 作為**獨立資料庫**儲存在 `annotations_cache.json`，retrieval 後把同 chunk 的註解一起帶下去：
# 
# | 階段 | 用法 |
# |---|---|
# | BM25 / chunk vector | **不變** — 註解不混進主索引，避免噪訊 |
# | BGE-reranker (步驟 10) | cross-encoder 看到 `chunk原文 + 【註解】` → 標準化關鍵字 (canonical_category, key_specs) 提升相關性命中 |
# | LLM (步驟 10b) | context 每段加 `【註解】` 區塊，等於先給 LLM 看一份產品 cheat-sheet |
# 
# **註解結構** (固定 YAML，從原文逐項抽取，禁止幻覺)：
# ```yaml
# canonical_category: <步驟 4.5 同義詞表的正式名稱>
# model_codes: [<原文出現的型號>]
# key_specs: {wattage, color_temp, lumens, ip_rating, voltage}
# aliases_present: [原文出現的別名]
# use_cases: [原文明確提到的應用場景]
# one_line_summary: <≤40 字一句話摘要>
# ```
# 
# **獨立性**:
# - 儲存：`annotations_cache.json` (key = chunk index)
# - 生成：跑一次 LLM (gpt-4o-mini, 388 chunks)，之後從 cache 讀
# - retrieval 端只在拿到 top-k 後 **paired lookup**，不參與排序的 BM25/向量計算
# 

# In[ ]:


# Cell — 步驟 4.7: 產品註解 sidecar 資料庫
# 對每個 chunk 用 LLM 生成 YAML 註解，cache 到 annotations_cache.json
# Retrieval 後 paired lookup: chunk_id → annotation_text

print("\n" + "=" * 60)
print("步驟 4.7: 產品註解獨立資料庫 (annotation sidecar DB)")
print("=" * 60)

import hashlib

ANNOTATIONS_CACHE_PATH = "./annotations_cache.json"
ANNOTATION_LLM = "gpt-4o-mini"  # 註解生成用便宜模型；388 chunks × 1 call
ANNOTATION_MAX_CHARS = 1500     # 截斷長 chunk，避免 prompt 爆掉

_canonical_list = sorted({c for g in SYN_GROUPS_LIST for c in g if c in ALIAS_TO_CANON.values()})
_alias_table_str = "\n".join(
    f"  - {canon}: {sorted(g - {canon})}"
    for canon, g in zip(
        [list(g)[0] for g in SYN_GROUPS_LIST],
        SYN_GROUPS_LIST,
    )
)

ANNOTATION_PROMPT_TMPL = """你是舞光 (Dancelight) LED 型錄資料標註員。任務：為下方「型錄節錄」產生一段結構化 YAML 註解，
讓後續的 RAG 系統 (BM25 + BGE-M3 + reranker + LLM) 更準確地把客戶問題對映到正確產品。

【嚴格規則】
1. 只能根據型錄節錄中**實際出現的文字**填寫，不得自行推論或新增規格。
2. 找不到的欄位填「未提供」，禁止留白或編造。
3. 產品類別 (canonical_category) 從下列正式名稱選 1 個；找不到對應則填「其他」：
   {canonical_list}
4. aliases_present 只能列出**節錄文字確實出現過**的別名 (從下表挑)：
   {alias_table}
5. 全部用繁體中文；YAML 區塊整體 ≤ 200 字。

【輸出格式】(只輸出純 YAML，不要任何前後說明)
canonical_category: <步驟 3 的正式名稱>
model_codes: [<型號1>, <型號2>, ...]
key_specs:
  wattage: <如 "30W" 或 "未提供">
  color_temp: <如 "3000K/4000K/6500K" 或 "未提供">
  lumens: <如 "3000lm" 或 "未提供">
  ip_rating: <如 "IP65" 或 "未提供">
  voltage: <如 "AC100-240V" 或 "未提供">
aliases_present: [<節錄出現的別名>]
use_cases: [<節錄明確提到的應用場景，最多 3>]
one_line_summary: <≤40 字一句話描述「這是哪一類、哪個型號、給誰用」>

【型錄節錄】
{chunk_text}

【註解輸出】
"""

def _hash_chunk(text):
    return hashlib.md5(text.encode("utf-8")).hexdigest()[:16]

def _generate_annotation(chunk_text):
    txt = chunk_text[:ANNOTATION_MAX_CHARS]
    prompt = ANNOTATION_PROMPT_TMPL.format(
        canonical_list=", ".join(_canonical_list),
        alias_table=_alias_table_str,
        chunk_text=txt,
    )
    try:
        resp = openai_client.chat.completions.create(
            model=ANNOTATION_LLM,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=800,
            temperature=0,
        )
        return resp.choices[0].message.content.strip()
    except Exception as e:
        return f"# annotation generation failed: {type(e).__name__}: {e}"

# Load cache
if os.path.exists(ANNOTATIONS_CACHE_PATH):
    with open(ANNOTATIONS_CACHE_PATH, "r", encoding="utf-8") as f:
        ANNOTATION_DB = json.load(f)
    print(f"  [cache] 載入 {len(ANNOTATION_DB)} 筆已生成註解")
else:
    ANNOTATION_DB = {}
    print(f"  [cache] 不存在，從零開始")

# Generate (skip cached)
to_generate = []
for idx, chunk_text in enumerate(chunks):
    h = _hash_chunk(chunk_text)
    if h not in ANNOTATION_DB:
        to_generate.append((idx, h, chunk_text))

print(f"  待生成 {len(to_generate)}/{len(chunks)} 筆 (LLM={ANNOTATION_LLM})")

if to_generate:
    t0 = time.time()
    for i, (idx, h, txt) in enumerate(to_generate):
        ann = _generate_annotation(txt)
        ANNOTATION_DB[h] = ann
        if (i + 1) % 25 == 0 or (i + 1) == len(to_generate):
            elapsed = time.time() - t0
            print(f"    {i+1}/{len(to_generate)} ({elapsed:.0f}s, "
                  f"avg {elapsed/(i+1):.1f}s/chunk)")
            # Periodic cache flush
            with open(ANNOTATIONS_CACHE_PATH, "w", encoding="utf-8") as f:
                json.dump(ANNOTATION_DB, f, ensure_ascii=False, indent=2)
    with open(ANNOTATIONS_CACHE_PATH, "w", encoding="utf-8") as f:
        json.dump(ANNOTATION_DB, f, ensure_ascii=False, indent=2)
    print(f"  完成；cache 已寫入 {ANNOTATIONS_CACHE_PATH}")

# Build chunk_id → annotation lookup
def get_annotation(chunk_id_or_text):
    """根據 chunk index 或 chunk text 取得註解"""
    if isinstance(chunk_id_or_text, int):
        h = _hash_chunk(chunks[chunk_id_or_text])
    else:
        h = _hash_chunk(chunk_id_or_text)
    return ANNOTATION_DB.get(h, "")

# Quick stats
non_empty = sum(1 for k, v in ANNOTATION_DB.items() if v and not v.startswith("#"))
print(f"  ANNOTATION_DB: {len(ANNOTATION_DB)} 筆, 有效註解 {non_empty} 筆")

# Sample (first chunk)
if chunks:
    sample_ann = get_annotation(0)
    print(f"\n  [Sample] chunk[0] (p.{chunk_metas[0].get('page','?')}):")
    print("  " + "\n  ".join(sample_ann.splitlines()[:8]))


# ## 步驟 5：BM25 索引
# 
# 建立 BM25 詞袋模型用於關鍵字檢索。
# 
# **`lamp_terms` 自訂詞典**: 因為 jieba 預設不認識「投射燈」「節能標章」這類複合詞，會被拆成 `投射 / 燈`，導致搜尋失準。我們手動把 38 個燈具相關複合詞加進 jieba 詞典。
# 
# **Tokenize**: 對每個 chunk 用 jieba 切詞並去空白，餵進 `BM25Okapi`。
# 
# **為什麼還要 BM25**: BGE-M3 向量對「LED-T813N-ESR1」這種獨特型號其實沒那麼準（型號通常沒語意），但 BM25 對精確字串完美。混合用兩者各取所長。

# In[ ]:


# Cell — 步驟 5: BM25 索引 (jieba 詞典 ← LAMP_TERMS, 來自 category_synonyms.xlsx)
print("\n" + "=" * 60)
print("步驟 5: BM25 索引")
print("=" * 60)

# LAMP_TERMS 已在步驟 4.5 從 xlsx 載入 (含所有正式名+別名+專業名詞)
for t in LAMP_TERMS:
    jieba.add_word(t)

tokenized_chunks = [[t.strip() for t in jieba.cut(ch) if t.strip()] for ch in chunks]
bm25 = BM25Okapi(tokenized_chunks)
print(f"BM25 索引完成 (LAMP_TERMS jieba 詞典: {len(LAMP_TERMS)} 詞)")


# ## 步驟 6：BGE-M3 向量化（含快取）+ 載入 Reranker
# 
# **為什麼選 BGE-M3**:
# - 中文表現是開源 embedder 第一線
# - 1024 維輸出，計算成本適中
# - 支援 dense / sparse / multi-vector 三種模式 (這裡只用 dense)
# - 可在本地 GPU 跑，不用打 API
# 
# **快取機制**: 第一次跑會把 386 chunks × 1024 dim 的矩陣 (~1.5MB) 存到 `bge_m3_embeddings/chunk_embeddings.npy`。之後重跑直接 load，省 30+ 秒模型載入時間。
# 
# **截斷至 4000 字**: BGE-M3 max 8192 tokens；太長的 chunk 截到 4000 字（約 8000 tokens）保險。
# 
# **`normalize_embeddings=True`**: 直接做 L2-normalize，後續 cosine similarity 變成單純內積，省一次正規化計算。
# 
# **載入 Reranker**: 同步載入 `BAAI/bge-reranker-v2-m3` CrossEncoder。這個模型不會被快取（它每次需要 query+doc 對，不能預先計算），但只佔 GPU 不算貴。

# In[ ]:


# Cell 14 — BGE-M3 向量化所有 chunks (有 .npy 快取)；同步載入 BGE-reranker CrossEncoder
print("\n" + "=" * 60)
print("步驟 6: BGE-M3 向量化")
print("=" * 60)

t0 = time.time()
embed_cache_file = os.path.join(EMBED_DIR, "chunk_embeddings.npy")
os.makedirs(EMBED_DIR, exist_ok=True)

if os.path.exists(embed_cache_file):
    chunk_embeddings = np.load(embed_cache_file)
    if chunk_embeddings.shape[0] == len(chunks):
        print(f"載入快取 embeddings: {chunk_embeddings.shape}")
    else:
        print(f"快取不一致 ({chunk_embeddings.shape[0]} vs {len(chunks)})，重建")
        chunk_embeddings = None
else:
    chunk_embeddings = None

if chunk_embeddings is None:
    print(f"載入 embedding 模型 ({EMBED_MODEL})...")
    embed_model = SentenceTransformer(EMBED_MODEL, **EMBED_KWARGS)
    print(f"模型載入完成，開始向量化 {len(chunks)} 個 chunks...")

    # 分批 embed (避免 OOM)
    BATCH = 32
    all_embs = []
    for i in range(0, len(chunks), BATCH):
        batch = chunks[i:i+BATCH]
        # 截斷過長的 chunk (BGE-M3 max 8192 tokens)
        batch = [c[:4000] for c in batch]
        embs = embed_model.encode(batch, show_progress_bar=False, normalize_embeddings=True)
        all_embs.append(embs)
        print(f"  向量化: {min(i+BATCH, len(chunks))}/{len(chunks)}", flush=True)

    chunk_embeddings = np.vstack(all_embs)
    np.save(embed_cache_file, chunk_embeddings)
    print(f"向量化完成，shape={chunk_embeddings.shape}")
    del embed_model  # 釋放 GPU 記憶體
else:
    pass

print(f"BGE-M3 embeddings ready, 耗時 {time.time()-t0:.1f}s")

# 載入 BGE-reranker
print(f"\n載入 BGE-reranker...")
t0 = time.time()
reranker = CrossEncoder(RERANK_MODEL)
print(f"BGE-reranker 載入完成，耗時 {time.time()-t0:.1f}s")


# ## 步驟 7：HyDE + Query Decomposition + Metadata Filter 函式
# 
# 這個 cell 定義三個關鍵函式（不執行）：
# 
# ### `hyde_generate(query)` — Hypothetical Document Embedding
# 餵客戶查詢給 LLM，要它「假裝寫一段對應的型錄條目」。然後用這段假設條目去做向量檢索，而不是直接拿原查詢。原理：型錄文字風格 vs 客戶查詢風格差很大，假設條目能讓向量更貼近真正的目標 chunk。
# 
# > 目前實際的 pipeline 沒在用 HyDE — 因為 `expand_specs` + `add_synonyms` 已經能補足查詢→型錄風格差異，HyDE 帶來的成本（一次 LLM 呼叫）不划算。函式留著給未來實驗。
# 
# ### `decompose_query(query)` — 規格條件抽取
# 對客戶查詢做 regex，抽出結構化規格 dict：
# 
# | key | 範例 | 來源 regex |
# |---|---|---|
# | `category` | `泛光燈` | 產品類別字典 (cat_map) 比對 |
# | `max_wattage` | `50` | `≦50W` 或 `50W 以下` |
# | `color_temp` | `6000` | `6000K` |
# | `min_lumens` | `6000` | `≧6000lm` |
# | `ip_rating` | `65` | `IP65` |
# | `features` | `[節能標章, 防水]` | 關鍵字比對 |
# 
# `cat_map` 用「長詞優先」順序確保「吊管式吊燈」不會被「吊燈」搶先匹配。
# 
# ### `metadata_filter(specs, chunk_metas_list)` — 結構化過濾
# 依規格條件回傳合格 chunk 的 index 集合：
# 
# - **類別**: 用 `synonyms` 同義詞群比對（如 `泛光燈` 群 = `{泛光燈, 投射燈, 探照燈, 聚光燈}`）
# - **瓦數**: 寬鬆過濾，只要該 chunk 任一型號 ≤ `max_wattage × 1.2`（20% 容差）就保留；無瓦數資訊的 chunk **不排除**（避免 metadata 抽取漏網而誤殺）
# - **IP**: 同上，IP 等級 ≥ 客戶要求才保留
# 
# 回傳的 `valid` 集合用於 `hybrid_retrieve` 中的 1.3× / 0.7× 加分機制（不是硬過濾）。

# In[ ]:


# Cell — 步驟 7: hyde_generate / decompose_query (← ALIAS_TO_CANON) / metadata_filter (← NAME_TO_GROUP)
print("\n" + "=" * 60)
print("步驟 7: HyDE + Query Decomposition + Metadata Filter (用 xlsx 同義詞表)")
print("=" * 60)


def hyde_generate(query):
    """用 LLM 生成假設的理想型錄條目，用於向量檢索"""
    prompt = f"""寫一段舞光照明型錄產品條目。包含：產品系列名、型號、瓦數、色溫、光通量、IP、材質。100字。

需求：{query}

型錄條目："""
    try:
        resp = ollama.chat(model=LOCAL_MODEL, messages=[{"role": "user", "content": prompt}],
                          options={"temperature": 0.3, "num_ctx": 2048, "num_predict": 200})
        content = resp["message"]["content"].strip()
        content = re.sub(r'<think>[\s\S]*?</think>', '', content).strip()
        return content if content else query
    except Exception as e:
        print(f"    [HyDE] 失敗: {e}")
        return query


def decompose_query(query):
    """用 regex + xlsx ALIAS_TO_CANON 從查詢提取結構化規格條件"""
    specs = {}

    # 類別 — 長詞優先 (LAMP_TERMS 已長序排) 掃 ALIAS_TO_CANON
    for term in LAMP_TERMS:
        if term in ALIAS_TO_CANON and term in query:
            specs["category"] = ALIAS_TO_CANON[term]
            break

    # 瓦數 (≦50W → max_wattage=50)
    m = re.search(r'[≦<]\s*(\d+)\s*W', query)
    if m:
        specs["max_wattage"] = int(m.group(1))
    else:
        m = re.search(r'(\d+)\s*W.*以下', query)
        if m:
            specs["max_wattage"] = int(m.group(1))

    # 色溫
    m = re.search(r'(\d{3,5})\s*K', query)
    if m:
        specs["color_temp"] = int(m.group(1))

    # 光通量 (≧6000lm → min_lumens=6000)
    m = re.search(r'[≧>]\s*(\d+)\s*(?:lm|LM)', query, re.I)
    if m:
        specs["min_lumens"] = int(m.group(1))

    # IP 等級
    m = re.search(r'IP\s*(\d+)', query)
    if m:
        specs["ip_rating"] = int(m.group(1))

    # 特殊功能
    features = []
    for feat in ["節能標章", "感應", "調光", "防眩", "防水"]:
        if feat in query:
            features.append(feat)
    if features:
        specs["features"] = features

    return specs


def metadata_filter(specs, chunk_metas_list):
    """根據 specs 過濾 chunk；類別比對使用 xlsx NAME_TO_GROUP（雙向）"""
    if not specs:
        return set(range(len(chunk_metas_list)))

    valid = set(range(len(chunk_metas_list)))

    # 類別過濾 — 使用 xlsx 同義詞群（雙向）
    # 注意：未標 category 的 chunk 也納入 (與 v5 行為一致；它們的 mc=="" 透過 substring fallback 通過)
    cat = specs.get("category", "")
    if cat:
        cat_syns = NAME_TO_GROUP.get(cat, {cat})
        cat_indices = set()
        for i, m in enumerate(chunk_metas_list):
            mc = m.get("category", "")
            mc_syns = NAME_TO_GROUP.get(mc, {mc}) if mc else set()
            # 命中條件：
            #   (a) chunk 的類別在查詢同義詞群裡
            #   (b) 查詢類別在 chunk 同義詞群裡
            #   (c) 字串包含 (mc=="" 此處恆真，等同保留無類別 chunk → 與 v5 一致)
            if (mc and mc in cat_syns) or (mc_syns and cat in mc_syns) \
               or (mc in cat) or (cat in mc):
                cat_indices.add(i)
        if cat_indices:
            valid &= cat_indices

    # 瓦數過濾 (寬鬆: 20% 容差)
    max_w = specs.get("max_wattage")
    if max_w and isinstance(max_w, (int, float)):
        w_indices = set()
        for i, m in enumerate(chunk_metas_list):
            watts = m.get("wattages", "")
            if not watts:
                w_indices.add(i)
                continue
            try:
                w_vals = [int(w) for w in watts.split(",") if w.strip()]
                if any(w <= max_w * 1.2 for w in w_vals):
                    w_indices.add(i)
            except ValueError:
                w_indices.add(i)
        valid &= w_indices

    # IP 過濾
    ip = specs.get("ip_rating")
    if ip and isinstance(ip, (int, float)):
        ip_indices = set()
        for i, m in enumerate(chunk_metas_list):
            m_ip = m.get("ip_rating", "")
            if not m_ip:
                ip_indices.add(i)
                continue
            try:
                if int(m_ip) >= int(ip):
                    ip_indices.add(i)
            except ValueError:
                ip_indices.add(i)
        valid &= ip_indices

    return valid


print("decompose_query (← ALIAS_TO_CANON) 與 metadata_filter (← NAME_TO_GROUP) 建立完成")


# ## 步驟 8：Hybrid Retriever — BM25 + 向量融合
# 
# ### `SYNONYMS` 同義詞群（基於試算表 `category_synonyms.xlsx`）
# 每個 group 中所有詞互相認對方，例：
# 
# ```
# 泛光燈 group = {泛光燈, 投射燈, 探照燈, 聚光燈}
# 崁燈 group = {崁燈, 筒燈, 嵌燈, 下照燈, downlight}
# 工事燈 group = {工事燈, 吊管燈, 吊管式吊燈, 停車場燈}
# ```
# 
# ### `add_synonyms(text)`
# 對輸入文字做同義詞展開：若文字中出現群組關鍵字，把整個群組的詞都接到後面。用於：
# 1. **BM25 查詢**: 客戶說「投射燈」也能命中型錄裡的「泛光燈」chunk
# 2. **Reranker query**: 同上，讓 cross-encoder 有更多匹配機會
# 
# ### `hybrid_retrieve(query, bm25_query, top_k, valid_indices, ...)`
# 
# 核心檢索函式。流程：
# 
# ```
# 1. BM25 score
#    tokens = jieba.cut(bm25_query)
#    bm25_scores = bm25.get_scores(tokens)
#    bm25_norm = bm25_scores / bm25_scores.max()    # 0~1 正規化
# 
# 2. 向量 score
#    q_emb = bge_m3.encode(query)                   # already L2-normalized
#    cos_scores = chunk_embeddings @ q_emb          # cosine = dot product
#    vec_norm = (cos_scores - min) / (max - min)    # min-max 正規化
# 
# 3. 融合
#    hybrid = 0.5 * bm25_norm + 0.5 * vec_norm
# 
# 4. Metadata 加分
#    if valid_indices given:
#        hybrid[i] *= 1.3 if i in valid_indices else 0.7
# 
# 5. 回傳 top_k (預設 RETRIEVE_K=50)
# ```
# 
# ### 為什麼是 1.3 / 0.7
# 不是硬過濾。如果客戶說「投射燈」但 metadata 抽取錯把該頁標成「軌道燈」，硬過濾會直接把正解踢出 top-50。1.3/0.7 只是加 / 減權重，給機會讓向量分數高的 chunk 翻身。
# 
# ### `clean_for_xlsx(text)`
# 工具函式：移除 xlsx 不接受的非法控制字元（`\x00-\x08` 等），用於最後輸出。

# In[ ]:


# Cell — 步驟 8: add_synonyms (← xlsx NAME_TO_GROUP) + hybrid_retrieve (BM25+vec 雙端展開)
print("\n" + "=" * 60)
print("步驟 8: Hybrid Retriever (BM25+BGE-M3, 兩端皆同義詞展開)")
print("=" * 60)

# 同義詞群直接從 xlsx 載入的 NAME_TO_GROUP；不再硬編碼
SYNONYMS = NAME_TO_GROUP


def add_synonyms(text):
    """掃描 text 中出現的任一同義詞，把整個群追加到字尾"""
    extra = set()
    for kw, syns in SYNONYMS.items():
        if kw in text:
            extra.update(syns)
    return text + " " + " ".join(extra) if extra else text


_query_embed_model = None
def _get_query_embedder():
    global _query_embed_model
    if _query_embed_model is None:
        _query_embed_model = SentenceTransformer(EMBED_MODEL, **EMBED_KWARGS)
    return _query_embed_model


def hybrid_retrieve(query, bm25_query=None, top_k=TOP_K,
                    valid_indices=None, bm25_w=BM25_WEIGHT, vec_w=VECTOR_WEIGHT):
    """BM25 + BGE-M3 融合檢索；BM25 與 向量 query 兩端都會經 add_synonyms 展開"""
    if bm25_query is None:
        bm25_query = query

    # ▼ 兩端皆同義詞展開
    bm25_query_x = add_synonyms(bm25_query)
    vec_query_x  = add_synonyms(query)

    # BM25
    qt = list(jieba.cut(bm25_query_x))
    bm25_scores = bm25.get_scores(qt)
    bm25_max = bm25_scores.max() if bm25_scores.max() > 0 else 1
    bm25_norm = bm25_scores / bm25_max

    # BGE-M3 vector (cosine via dot product on normalized embeddings)
    model = _get_query_embedder()
    q_emb = model.encode([vec_query_x[:2000]], normalize_embeddings=True)
    cos_scores = (chunk_embeddings @ q_emb.T).flatten()
    vec_min, vec_max = cos_scores.min(), cos_scores.max()
    vec_norm = (cos_scores - vec_min) / (vec_max - vec_min + 1e-8)

    # 融合
    hybrid = bm25_w * bm25_norm + vec_w * vec_norm

    # metadata 過濾：符合者加分，不符合輕微降權 (1.3× / 0.9×)
    if valid_indices is not None and len(valid_indices) < len(chunks):
        for i in range(len(chunks)):
            if i in valid_indices:
                hybrid[i] *= 1.3
            else:
                hybrid[i] *= 0.9

    top_idx = np.argsort(hybrid)[::-1][:top_k]
    results = []
    for idx in top_idx:
        results.append({
            "chunk_id": int(idx),
            "text": chunks[idx],
            "metadata": chunk_metas[idx],
            "score": float(hybrid[idx]),
            "bm25_score": float(bm25_norm[idx]),
            "vector_score": float(vec_norm[idx]),
        })
    return results


print("Hybrid Retriever (BGE-M3 + 雙端同義詞展開) 建立完成")

_ILLEGAL_CHARS_RE = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]')
def clean_for_xlsx(text):
    if isinstance(text, str):
        return _ILLEGAL_CHARS_RE.sub('', text)
    return text

print("工具函式建立完成")


# ## 步驟 9：查詢規格擴展 `expand_specs`
# 
# 對客戶查詢做文字級的「單位語意化」改寫。例：
# 
# ```
# 原始: 產品，LED 投射燈，≦50W，6000K，≧6000lm，IP65
# 擴展: 產品，LED 投射燈，小於等於50瓦(W)，6000色溫(K)，大於等於6000光通量(lm)，IP65
# ```
# 
# **為什麼這樣做**: BGE-M3 向量看到「50W」可能聯想成「50 watts」「50 瓦」「50W LED」等，但**未必特別偏好型錄裡的「50瓦(W)」格式**。手動把單位數字 + 中文化讓 chunk 端 (型錄文字也類似格式) 與 query 端格式對齊，提高向量相似度。
# 
# **處理項目**:
# - `lm/W` → `發光效率(lm/W)`
# - `(\d+)W` → `(\d+)瓦(W)`
# - `(\d+)K` → `(\d+)色溫(K)`
# - `(\d+)lm` → `(\d+)光通量(lm)`
# - `(\d+)V` → `(\d+)電壓(V)`
# - `Ra(\d+)` → `演色性Ra(\d+)`
# - `≦` → `小於等於`，`≧` → `大於等於`

# In[ ]:


# Cell 20 — expand_specs() 把查詢中的 50W/6000K/≧6000lm 改寫成單位語意化文字
def expand_specs(text):
    t = text
    t = re.sub(r'(\d+)\s*lm/W', r'\1發光效率(lm/W)', t)
    t = re.sub(r'([A-Za-z])(\d)', r'\1 \2', t)
    t = re.sub(r'(\d+)\s*W(?![a-zA-Z/])', lambda m: m.group(1) + "瓦(W)", t)
    def expand_k(m):
        prefix = t[:m.start()]
        if prefix.endswith("色溫"):
            return m.group(1) + "K"
        return m.group(1) + "色溫(K)"
    t = re.sub(r'(\d+)\s*K(?![a-zA-Z])', expand_k, t)
    t = re.sub(r'(\d+)\s*lm(?![/a-zA-Z])', lambda m: m.group(1) + "光通量(lm)", t, flags=re.IGNORECASE)
    t = re.sub(r'(\d+)\s*V(?![a-zA-Z])', lambda m: m.group(1) + "電壓(V)", t)
    t = re.sub(r'Ra\s*(\d+)', r'演色性Ra\1', t)
    t = t.replace("≦", "小於等於").replace("≧", "大於等於")
    return t


# ## 步驟 10：Training 測試（完整 pipeline 整合）
# 
# 讀 `Training.xlsx` 的測試題庫（兩欄：詢問問題 / 期望回答），對每題跑完整檢索流程：
# 
# ```
# for each question:
#     ① decompose_query(original)            → specs dict
#     ② metadata_filter(specs, chunk_metas)  → valid_indices set
#     ③ expanded = expand_specs(original)    → 改寫後的 query
#        bm25_q   = add_synonyms(original)   → BM25 同義詞展開
#     ④ hybrid_retrieve(expanded, bm25_q,    → 50 candidates
#                       top_k=50, valid_indices)
#     ⑤ BGE-reranker:
#        pairs = [[add_synonyms(original), c.text[:1000]] for c in candidates]
#        scores = reranker.predict(pairs)
#        sort & take top 20
# ```
# 
# **為什麼 reranker 也要做 `add_synonyms`**: cross-encoder 是逐對匹配，若客戶查詢只有「投射燈」而正解 chunk 標的是「泛光燈」，cross-encoder 可能給低分。展開同義詞讓 query 端有「泛光燈」字樣，匹配機會增加。
# 
# **Top-1000 字截斷**: BGE-reranker 對長 chunk 處理慢；前 1000 字通常已含結構化摘要 + 大部分原文。
# 
# **結果儲存**: `results: list[dict]` 每筆含 `qid / original / expanded / expected / specs / docs(top-20)`，傳給下一個 cell 寫 xlsx。

# In[ ]:


# Cell 22 — 主測試迴圈：對 Training.xlsx 每題跑 decompose→filter→retrieve→rerank 完整 pipeline
print("\n" + "=" * 60)
print("步驟 10: Training 資料測試")
print("  Pipeline: Query Decomposition → Metadata Filter")
print("            → HyDE → Hybrid Retrieve → LLM Rerank")
print("=" * 60)

df_q = pd.read_excel(QUESTION_XLSX)
df_a = pd.read_excel(ANSWER_XLSX)
questions = df_q.iloc[:, 0].astype(str).tolist()
answers   = df_a.iloc[:, 0].astype(str).tolist()
assert len(questions) == len(answers), \
    f"question/answer 行數不符: {len(questions)} vs {len(answers)}"
print(f"測試題數: {len(questions)}")

results = []

for idx, (original, expected) in enumerate(zip(questions, answers)):
    qid = idx
    original = original.strip()
    expected = expected.strip()

    print(f"\n{'='*50}")
    print(f"Q{qid}: {original[:60]}")
    print(f"期望: {expected[:50]}")

    # Step 1: Query Decomposition (regex, 瞬間)
    specs = decompose_query(original)
    print(f"  [Specs] {specs}")

    # Step 2: Metadata Filter
    valid_indices = metadata_filter(specs, chunk_metas)
    print(f"  [Filter] {len(valid_indices)}/{len(chunks)} chunks 加分")

    # Step 3: Hybrid Retrieve  (hybrid_retrieve 內部會對 BM25 與 向量 query 兩端做 add_synonyms)
    t0 = time.time()
    expanded = expand_specs(original)
    candidates = hybrid_retrieve(expanded, bm25_query=original,
                                  top_k=RETRIEVE_K, valid_indices=valid_indices)
    t_retrieve = time.time() - t0

    # Step 4: BGE-reranker (Cross-Encoder) — 用同義詞展開的查詢 + chunk 註解
    t1 = time.time()
    rerank_query = add_synonyms(original)
    # paired lookup: 每個候選 chunk 同時帶上其 sidecar 註解
    for c in candidates:
        c["annotation"] = get_annotation(c["chunk_id"])
    def _augment(c):
        ann = c["annotation"]
        body = c["text"][:800]
        return body + ("\n\n【註解】\n" + ann[:600] if ann else "")
    pairs = [[rerank_query, _augment(c)] for c in candidates]
    rerank_scores = reranker.predict(pairs)
    for d, rs in zip(candidates, rerank_scores):
        d["rerank_score"] = float(rs)
        d["rerank_reason"] = ""
    # 按 rerank 分數排序
    candidates.sort(key=lambda x: x["rerank_score"], reverse=True)
    candidates = candidates[:TOP_K]
    t_rerank = time.time() - t1

    print(f"  [Retrieve] {t_retrieve:.1f}s | [Rerank] {t_rerank:.1f}s")
    for i, d in enumerate(candidates[:5]):
        pg = d["metadata"].get("page", "?")
        cat = d["metadata"].get("category", "")
        rs = d["rerank_score"]
        hs = d["score"]
        print(f"  #{i+1} [p.{pg}] rerank={rs:.4f} hybrid={hs:.3f} cat={cat}")

    results.append({
        "qid": qid,
        "original": original,
        "expanded": expanded,
        "expected": expected,
        "specs": specs,
        "docs": candidates,
    })


# ## 步驟 10b：LLM 答案生成（5 選項格式）
# 
# 把 reranker 排序後的 top-20 chunks 餵給本地 LLM (`gpt-oss:120b`) 產生最終回覆。輸出格式強制為：
# 
# ```
# ★ 推薦：<產品型號> — <瓦數/色溫/關鍵規格>
# 備選 1：<產品型號> — <規格>
# 備選 2：<產品型號> — <規格>
# 備選 3：<產品型號> — <規格>
# 備選 4：<產品型號> — <規格>
# ```
# 
# ### 為什麼是 5 選項而不是直接給答案
# 
# 1. **推薦 + 備選**: 強迫 LLM 從 20 個候選收斂到 1 個推薦 + 4 個備選，避免它隨便選一個。
# 2. **可雙重評估**: 後續 `score_llm` 同時計算
#    - **acc_any** — 期望型號是否出現在 5 選項中任一位 (寬鬆)
#    - **acc_rec** — 期望型號是否出現在 ★ 推薦那行 (嚴格)
# 3. **緩衝**: 真實場景下，當業務看到推薦不合心意，4 個備選提供 fallback。
# 
# ### 為什麼用本地 LLM
# 
# 維持與整體 BGE-M3 + reranker 全部離線的一致性。如要改雲端 (OpenAI gpt-5 / Gemini)，把 `run_llm` 內部換掉即可，介面不動。
# 
# ### Context 截斷
# 
# `build_context()` 把 top-N 串接成單一 context，每段加 `[文件 N / p.頁碼]` 標頭，總字數限 `CTX_MAX_CHARS=12000`。原因：
# - gpt-oss:120b 預設 `num_ctx=2048` 太短；放寬到 8192 是安全上限
# - 超過 12000 字的 chunk 通常是冗餘文字，截斷對命中率影響很小，但能顯著降低 latency
# 
# ### 輸出處理
# 
# `<think>...</think>` 段是某些 reasoning 模型的內部思考過程，會被 regex 移除以保留純答案。
# 
# ### v3.2 (2026-05-13) 變更說明
# 
# 多模態實驗已撤回 (見 cell 0 說明)。本 cell 程式碼維持 v3.1 純文字輸入版本。

# In[ ]:


# Cell 24 — 用 OpenAI gpt-4o (雲端) 從 top-20 chunks 產生 5 選項回答
# 改為雲端 LLM 因為本地 gpt-oss/qwen3.6/gemma4 在長中文 prompt 下會幻覺或回空
print("\n" + "=" * 60)
print(f"步驟 10b: LLM 答案生成 ({CLOUD_LLM})")
print("=" * 60)

# v3 fix: 把 char budget 提高到能容納全部 20 chunks (約 40k 字 = ~12k tokens)
# 雲端 LLM 有 128k+ context，不再受本地模型 16k limit 限制
CTX_MAX_CHARS = 40000
PER_DOC_MAX = 2000

RAG_PROMPT_TMPL = '''你是舞光 (Dancelight) LED 產品顧問。從下列型錄節錄中為客戶推薦【5 個最接近的產品】，第 1 個標示為「★ 推薦」，其餘 4 個為「備選方案」。

**重要原則**：
- **產品型號必須完全照型錄節錄原文** (含大小寫、橫線、字尾編號)，不要自己生成或修改。
- 即使規格不完全吻合，也要從節錄中**選最接近的 5 個型號**列出。客戶要的是建議清單，不是硬性篩選。
- 只有在型錄節錄中**完全沒有相同類別**的產品時，才回覆「無匹配產品」。其他情況都要列出 5 個候選。

輸出格式（務必嚴格遵守）：
★ 推薦：<型號> — <瓦數/色溫/關鍵規格>
備選 1：<型號> — <規格>
備選 2：<型號> — <規格>
備選 3：<型號> — <規格>
備選 4：<型號> — <規格>

=== 型錄節錄 ===
{context}

=== 使用者問題 ===
{query}

=== 回答（5 個選項，第 1 個為推薦） ==='''


def build_context(docs, max_chars=CTX_MAX_CHARS, per_doc=PER_DOC_MAX):
    """把 top-K docs 串成 context；每段加 [文件 N / p.頁碼] 標頭。
    若 chunk 帶有 sidecar 註解 (annotation_db)，先把【註解】prepend 到原文前。"""
    parts, total = [], 0
    for i, d in enumerate(docs):
        head = f'[文件 {i+1} / p.{d["metadata"].get("page","?")}]\n'
        ann = d.get("annotation", "")
        ann_block = f"【註解】\n{ann[:600]}\n" if ann and not ann.startswith("#") else ""
        body = "【原文】\n" + d["text"][:per_doc] if ann_block else d["text"][:per_doc]
        chunk = head + ann_block + body
        if total + len(chunk) > max_chars:
            break
        parts.append(chunk)
        total += len(chunk)
    return "\n\n".join(parts), len(parts)


import requests as _requests
_OLLAMA_URL = "http://localhost:11434/api/chat"
_THINK_TAG_RE = re.compile(r"<think>.*?</think>", re.DOTALL | re.IGNORECASE)

# v3.3a: 抑制 qwen3 思考模式的三重保險
#   1) Qwen3 識別系統訊息中的 /no_think 指令 (官方支援)
#   2) think=False top-level + options.think=False (Ollama 雙端)
#   3) num_predict 大幅放寬，即使思考漏出來也有空間給格式化答案
SYS_NO_THINK = "/no_think\n你只能直接輸出最終答案，禁止輸出任何分析、思考、推理過程。"
NUM_PREDICT  = 12000   # v3.3c: 12k 預算 (折衷; 上次 6k 不夠、120k 太離譜)
NUM_CTX      = 32768   # 12k output + ~6k input prompt 還有餘裕

def run_llm(query, docs):
    """v3.3a: Ollama 本地 qwen3.6:latest，多重 no-think + 放寬 token budget + <think> 標籤剝除。"""
    context, n_docs = build_context(docs)
    prompt = RAG_PROMPT_TMPL.format(context=context, query=query)
    t0 = time.time()
    try:
        r = _requests.post(_OLLAMA_URL, json={
            "model": CLOUD_LLM,
            "messages": [
                {"role": "system", "content": SYS_NO_THINK},
                {"role": "user",   "content": prompt},
            ],
            "stream": False,
            "think": False,
            "options": {
                "num_predict": NUM_PREDICT,
                "temperature": 0,
                "num_ctx": NUM_CTX,
                "think": False,
            },
        }, timeout=1800)
        d = r.json()
        msg = d.get("message", {}) or {}
        ans = (msg.get("content") or "").strip()
        if not ans:
            # 部分版本把答案塞到 thinking/reasoning，照樣取回再過濾
            ans = (msg.get("thinking") or msg.get("reasoning") or "").strip()
        # 萬一還是有 <think>…</think> 區塊，剝除後再回傳
        ans = _THINK_TAG_RE.sub("", ans).strip()
        return {"answer": ans, "time": time.time() - t0, "error": "", "n_docs_in_context": n_docs}
    except Exception as e:
        return {"answer": "", "time": time.time() - t0, "error": f"{type(e).__name__}: {e}", "n_docs_in_context": n_docs}


for r in results:
    print(f"\n--- Q{r['qid']:02d}: {r['original'][:50]} ---")
    res = run_llm(r['original'], r['docs'])
    r['answer'] = res['answer']
    r['llm_time'] = res['time']
    r['llm_error'] = res['error']
    if res['error']:
        print(f"  ERROR: {res['error']}")
    else:
        first_lines = '\n  '.join(res['answer'].split('\n')[:30])
        print(f"  ({res['time']:.1f}s, {res['n_docs_in_context']}/{len(r['docs'])} docs in context)\n  {first_lines}")

print(f"\nLLM 答案生成完成")


# ## 步驟 11：xlsx 結果輸出
# 
# 把 `results` 寫成 Excel，**最終解果以 LLM 輸出為主**（reranker 結果作為佐證資料留在每題 sheet 裡）。
# 
# ### Summary sheet（給業務/老闆看的「答案總表」）
# 每題一列，欄位：
# 
# | 欄位 | 內容 |
# |---|---|
# | 題號 | qid |
# | 原始查詢 | 客戶提問原文 |
# | 期望回答 | 標準答案 (gold) |
# | ★ 推薦 | LLM 5 選項中第一行的「★ 推薦」內容 |
# | 完整 LLM 答案 | 整段 5 選項回覆 (★ 推薦 + 4 備選) |
# | 命中狀態 | HIT / MISS / SKIP，含實際命中型號 |
# 
# ### 每題 sheet (Q00, Q01, ...) — 證據與檢索診斷
# 每題一個分頁，內容由上而下：
# 
# 1. **題號 + 原始查詢**（粗體標頭）
# 2. **期望答案**（紅色粗體，方便比對）
# 3. **Specs dict**（decompose_query 抽出的結構化條件）
# 4. **★ LLM 5 選項回覆**（橘色標頭，整段 LLM 答案）
# 5. **檢索證據表 (top-20 reranker 結果)**：序號 / 頁碼 / 類別 / Rerank 分數 / 內容預覽 (前 400 字)
#    ─ 用於檢查 LLM 是否從這 20 段裡選對東西
# 
# ### 樣式
# - 標頭藍底白字粗體
# - LLM 答案區塊橘底白字
# - 全表細邊框、自動換行
# - `clean_for_xlsx` 過濾掉 Excel 不接受的控制字元 (`\x00-\x08` 等)

# In[ ]:


# Cell 26 — 把 results 寫成 rag_results.xlsx (Summary 顯示 LLM 答案 + 每題證據 sheet)
print("\n" + "=" * 60)
print("步驟 11: xlsx 輸出 (LLM 最終答案 + reranker 證據)")
print("=" * 60)

wb = Workbook()
ws = wb.active
ws.title = "Summary"

# ── 樣式 ──
hf = Font(bold=True, color="FFFFFF")
hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
ofill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
tb = Border(left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"))
wa = Alignment(wrap_text=True, vertical="top")


def extract_recommended(text):
    """從 LLM 答案抓 ★ 推薦 那行 (任何包含 ★/推薦/recommend 的行)"""
    if not text:
        return ""
    for line in text.split("\n"):
        if any(k in line.lower() for k in ["★", "推薦", "recommend"]):
            return line.strip()
    # fallback: first non-empty line
    for line in text.split("\n"):
        if line.strip():
            return line.strip()
    return ""


def hit_status(expected, answer, docs):
    """用 score_llm + check_hit 評估狀態 (cell 28 的函式必須先執行；若還沒，做 fallback)"""
    try:
        s = score_llm(expected, answer)
        if expected.strip() == "無匹配產品" or not extract_model_numbers(expected):
            return "SKIP (無匹配產品)"
        if s["hit_any"] > 0:
            return f"HIT [{','.join(s['hits_any'])}]"
        # 看檢索層是否命中
        retr_hit, hits, misses = check_hit(expected, docs)
        if retr_hit:
            return f"LLM-MISS (retr-HIT: {hits})"
        return f"MISS (期望: {misses})"
    except NameError:
        return "(尚未跑命中率統計)"


# ── Summary sheet ──
headers = ["題號", "原始查詢", "期望回答", "★ 推薦", "完整 LLM 答案", "命中狀態"]
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = hf; cell.fill = hfill; cell.border = tb; cell.alignment = wa

for r in results:
    answer = r.get("answer", "") or ""
    rec = extract_recommended(answer)
    status = hit_status(r["expected"], answer, r.get("docs", []))
    row = [
        r["qid"],
        r["original"],
        r["expected"],
        clean_for_xlsx(rec),
        clean_for_xlsx(answer),
        status,
    ]
    ri = r["qid"] + 2 if isinstance(r["qid"], int) else len(ws.iter_rows()) + 1
    for c, v in enumerate(row, 1):
        cell = ws.cell(row=ri, column=c, value=v)
        cell.border = tb; cell.alignment = wa

ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 45
ws.column_dimensions["C"].width = 30
ws.column_dimensions["D"].width = 50    # ★ 推薦
ws.column_dimensions["E"].width = 80    # 完整 LLM 答案
ws.column_dimensions["F"].width = 30    # 命中狀態

# ── 每題 sheet (Q00, Q01, ...) ──
for r in results:
    qid = r["qid"]
    sheet_title = f"Q{qid:02d}" if isinstance(qid, int) else f"Q{qid}"
    wq = wb.create_sheet(title=sheet_title)
    answer = r.get("answer", "") or ""

    # 1. 標頭
    wq.merge_cells("A1:F1")
    wq["A1"] = f"{sheet_title}: {r['original'][:80]}"
    wq["A1"].font = Font(bold=True, size=12)

    # 2. 期望答案 (紅色)
    wq["A2"] = "期望:"
    wq["A2"].font = Font(bold=True, color="FF0000")
    wq.merge_cells("B2:F2")
    wq["B2"] = clean_for_xlsx(r["expected"]); wq["B2"].alignment = wa

    # 3. Specs
    wq["A3"] = "Specs:"; wq["A3"].font = Font(bold=True)
    wq.merge_cells("B3:F3")
    wq["B3"] = str(r.get("specs", {})); wq["B3"].alignment = wa

    # 4. LLM 答案區 (橘色標頭)
    wq.merge_cells("A4:F4")
    wq["A4"] = "★ LLM 5 選項回覆"
    wq["A4"].font = hf; wq["A4"].fill = ofill; wq["A4"].alignment = Alignment(horizontal="center")
    wq.merge_cells("A5:F5")
    wq["A5"] = clean_for_xlsx(answer)
    wq["A5"].alignment = wa
    wq.row_dimensions[5].height = 180

    # 5. 命中狀態 + LLM latency
    wq.merge_cells("A6:F6")
    status = hit_status(r["expected"], answer, r.get("docs", []))
    lat = r.get("llm_time", 0)
    err = r.get("llm_error", "")
    wq["A6"] = f"狀態: {status}  |  LLM 耗時: {lat:.1f}s" + (f"  |  ERR: {err}" if err else "")
    wq["A6"].font = Font(bold=True)
    wq["A6"].alignment = wa

    # 6. 檢索證據表 (top-20)
    wq.merge_cells("A7:F7")
    wq["A7"] = "檢索證據 (top-20 reranker 結果)"
    wq["A7"].font = hf; wq["A7"].fill = hfill; wq["A7"].alignment = Alignment(horizontal="center")

    hdr_row = 8
    for c, h in enumerate(["#", "頁碼", "類別", "Rerank", "Hybrid", "內容預覽"], 1):
        cell = wq.cell(row=hdr_row, column=c, value=h)
        cell.font = hf; cell.fill = hfill; cell.border = tb; cell.alignment = wa

    for i, d in enumerate(r.get("docs", [])):
        ri = hdr_row + 1 + i
        wq.cell(row=ri, column=1, value=i + 1).border = tb
        wq.cell(row=ri, column=2, value=d["metadata"].get("page", "?")).border = tb
        wq.cell(row=ri, column=3, value=d["metadata"].get("category", "")).border = tb
        wq.cell(row=ri, column=4, value=round(d.get("rerank_score", 0), 4)).border = tb
        wq.cell(row=ri, column=5, value=round(d.get("score", 0), 3)).border = tb
        cell = wq.cell(row=ri, column=6, value=clean_for_xlsx(d["text"][:400]))
        cell.border = tb; cell.alignment = wa

    wq.column_dimensions["A"].width = 4
    wq.column_dimensions["B"].width = 6
    wq.column_dimensions["C"].width = 10
    wq.column_dimensions["D"].width = 10
    wq.column_dimensions["E"].width = 10
    wq.column_dimensions["F"].width = 100

wb.save(RESULTS_XLSX)
print(f"已儲存: {RESULTS_XLSX}")
print(f"  - Summary sheet: {len(results)} rows (LLM 答案為主)")
print(f"  - 每題 sheet: Q00 ~ Q{len(results)-1:02d}")


# ## 步驟 12：命中率統計
# 
# 定義 `check_hit(expected, docs, top_n=20)` 並對全部結果計算命中率。
# 
# ### `check_hit` 邏輯
# 1. 從「期望回答」欄解析出產品型號：
#    - `LED-4140R5+LED-T813N-ESR1` → `[LED-4140R5, LED-T813N-ESR1]`
#    - 用 regex `^[A-Za-z0-9\-]+` 抓開頭的型號字串，忽略後面中文說明
#    - 「無匹配產品」→ 跳過該題不計分
# 2. 把 top-20 chunk 的文字串接成一大段
# 3. 期望型號**任一**出現在文字中 → HIT；全部沒出現 → MISS
# 
# ### 統計輸出
# ```
# Q1: HIT [LED-4140R5] (未找到: ['LED-T813N-ESR1'])
# Q2: SKIP (無匹配產品)
# Q3: MISS (期望: ['D-21DOP25NR2'])
# ...
# 檢索命中率: 7/15 = 46.7%
# (跳過無匹配產品: 2 題)
# ```
# 
# 「任一型號命中即 HIT」是寬鬆評估 — 真實場景多型號組合（如「燈具+燈管」搭配）需要全部命中才算真正解，但當前指標已足夠用於迭代調參。

# In[ ]:


# Cell 28 — check_hit (retrieval) + score_llm (any-of-5 / recommended) + 整體命中率
print("\n" + "=" * 60)
print("步驟 12: 命中率統計 (檢索 + LLM)")
print("=" * 60)

# Chinese↔Latin safe model-number regex (no \b — CJK chars are word chars)
MODEL_NUM_RE = re.compile(
    r'(?<![A-Z0-9-])[A-Z][A-Z0-9]*(?:-[A-Z0-9]+)+(?![A-Z0-9-])'
    r'|(?<![A-Z0-9-])[A-Z]+\d[A-Z0-9]*(?![A-Z0-9-])'
)
UNIT_RE = re.compile(r'^(?:IP\d+|T\d+|CNS|CE|RoHS|LED)$', re.I)
REC_LINE_RE = re.compile(r'(?:★|推薦|recommend)[^\n]*', re.IGNORECASE)


def extract_model_numbers(text):
    """從文字抽出像 LED-4140R5 / D-21DOP25NR2 / E-FLCS50D / L4140R5 這樣的型號。
    支援 + 連接 (LED-2441R1+D-T810DR9) 和「搭配」連接 (LED-4140R5 搭配LED-T813N-ESR1)。"""
    if not text:
        return []
    out, seen = [], set()
    for m in MODEL_NUM_RE.findall(text):
        if m in seen:
            continue
        if UNIT_RE.match(m):  # 跳過 IP65 / T8 / CNS / LED 等非型號
            continue
        if len(m) < 5 and '-' not in m:
            continue
        out.append(m)
        seen.add(m)
    return out


def extract_recommended_line(text):
    """抓 LLM 答案中的 ★ 推薦 那行"""
    if not text:
        return ''
    m = REC_LINE_RE.search(text)
    return m.group(0) if m else ''


def check_hit(expected, docs, top_n=20):
    """檢索層命中率: 期望型號是否出現在 top-N chunk 文字裡"""
    expected = expected.strip()
    if expected == "無匹配產品":
        return None, [], []
    models = extract_model_numbers(expected)
    if not models:
        return None, [], []  # 無可解析型號 (如 q04 的 free-text 回退說明)
    combined = "\n".join(d["text"] for d in docs[:top_n]).upper()
    hits = [m for m in models if m.upper() in combined]
    misses = [m for m in models if m.upper() not in combined]
    return len(hits) > 0, hits, misses


def score_llm(expected, answer):
    """LLM 層命中率: 期望型號是否在 5 選項 (any) / 在 ★ 推薦 (rec)"""
    models = extract_model_numbers(expected)
    if not models:
        return {"hit_any": 0, "hit_rec": 0, "total": 0, "hits_any": [], "hits_rec": []}
    ans_upper = (answer or '').upper()
    rec_upper = extract_recommended_line(answer or '').upper()
    hits_any = [m for m in models if m.upper() in ans_upper]
    hits_rec = [m for m in models if m.upper() in rec_upper]
    return {
        "hit_any": len(hits_any), "hit_rec": len(hits_rec),
        "total": len(models), "hits_any": hits_any, "hits_rec": hits_rec,
    }


# 三層彙總: 檢索 (per question) + LLM (per code: hit_any / hit_rec)
hit_count = miss_count = skip_count = 0
llm_any_hit = llm_rec_hit = llm_total_codes = 0

for r in results:
    # 檢索層
    result, hits, misses = check_hit(r["expected"], r["docs"])
    if result is None:
        skip_count += 1
        retr_status = "SKIP (無匹配產品 / 無型號)"
    elif result:
        hit_count += 1
        retr_status = f"HIT [{','.join(hits)}]" + (f" (未找到: {misses})" if misses else "")
    else:
        miss_count += 1
        retr_status = f"MISS (期望: {misses})"

    # LLM 層
    s = score_llm(r["expected"], r.get("answer", ""))
    llm_any_hit += s["hit_any"]
    llm_rec_hit += s["hit_rec"]
    llm_total_codes += s["total"]
    llm_status = f"LLM any={s['hit_any']}/{s['total']} rec={s['hit_rec']}/{s['total']}"

    print(f"  Q{r['qid']:02d}: {retr_status}  |  {llm_status}")

# 彙總
total_scored = hit_count + miss_count
print()
print(f"題數:                          {len(results)} (跳過 {skip_count} 題)")
if total_scored > 0:
    print(f"檢索命中率 (gold in top-20):  {hit_count}/{total_scored} = {hit_count/total_scored*100:.1f}%")
if llm_total_codes > 0:
    print(f"LLM 任一選項命中 (acc_any):    {llm_any_hit}/{llm_total_codes} = {llm_any_hit/llm_total_codes*100:.1f}%")
    print(f"LLM 推薦命中     (acc_rec):    {llm_rec_hit}/{llm_total_codes} = {llm_rec_hit/llm_total_codes*100:.1f}%")

print("\n" + "=" * 60)
print("完成！請檢查 rag_results.xlsx")
print("=" * 60)


# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:




